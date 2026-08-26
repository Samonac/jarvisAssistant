"""Flask route definitions for Jarvis Assistant.

Defines all HTTP endpoints:
- POST /chat: Conversational endpoint
- GET /health: Health check
- GET/POST /login, POST /logout: Authentication
- GET /, /dashboard, /settings: Frontend routes (protected)
- GET /api/config, PUT /api/config: Configuration management (protected)
- GET /api/metrics: Usage metrics (protected)
- GET /api/system: System resource monitoring (protected)
"""

import uuid
import logging

from flask import Flask, request, jsonify, render_template, redirect, session

logger = logging.getLogger(__name__)


def register_routes(app: Flask) -> None:
    """Register all route handlers on the Flask app."""
    from app.iam import PERMISSION_GROUPS

    @app.context_processor
    def inject_user_permissions():
        """Make user permissions available in all templates."""
        iam = app.config.get("IAM_MANAGER")
        username = session.get("username", "")
        if iam and username:
            perms = iam.get_user_permissions(username)
        else:
            perms = list(PERMISSION_GROUPS.keys())  # No IAM = full access
        role = session.get("role", "")
        return {"user_permissions": perms, "current_user": username, "user_role": role}

    def _get_auth_manager():
        return app.config.get("AUTH_MANAGER")

    def _get_current_username():
        """Get the current authenticated username from session or Basic Auth."""
        # Check session first (web UI)
        if session.get("authenticated") and session.get("username"):
            return session.get("username")
        # Check Basic Auth header (API calls)
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Basic "):
            import base64
            try:
                decoded = base64.b64decode(auth_header[6:]).decode("utf-8")
                username, password = decoded.split(":", 1)
                iam = app.config.get("IAM_MANAGER")
                if iam:
                    user = iam.authenticate(username, password)
                    if user:
                        return username
            except (ValueError, UnicodeDecodeError):
                pass
        return ""

    def _get_user_role():
        """Get the current user's role from the session or Basic Auth."""
        username = _get_current_username()
        if not username:
            return ""
        iam = app.config.get("IAM_MANAGER")
        if iam and username in iam.users:
            return iam.users[username].get("role", "user")
        return session.get("role", "user")

    def _get_user_permissions():
        """Get the current user's permission list."""
        username = _get_current_username()
        iam = app.config.get("IAM_MANAGER")
        if iam and username:
            return iam.get_user_permissions(username)
        # Legacy: if no IAM or no user, check session
        if session.get("authenticated"):
            return list(PERMISSION_GROUPS.keys())
        # Auth entirely disabled (no WEB_PASSWORD configured) grants full
        # access, consistent with AuthManager.is_authenticated()'s own bypass.
        auth = _get_auth_manager()
        if auth and not auth.enabled:
            return list(PERMISSION_GROUPS.keys())
        return []

    def _is_authenticated():
        """Check if the current request is authenticated (session or Basic Auth)."""
        # Session auth (web UI)
        auth = _get_auth_manager()
        if auth and auth.is_authenticated():
            return True
        # Basic Auth (API)
        if _get_current_username():
            return True
        return False

    def _login_required(f):
        """Wrapper that checks authentication via session or Basic Auth."""
        import functools

        @functools.wraps(f)
        def decorated(*args, **kwargs):
            if not _is_authenticated():
                if request.path.startswith("/api/"):
                    return jsonify({"error": "Authentication required. Use session cookie or Basic Auth (Authorization: Basic base64(user:pass))"}), 401
                return redirect("/login")
            return f(*args, **kwargs)

        return decorated

    def _permission_required(permission):
        """Decorator factory: requires the user to have a specific permission."""
        def decorator(f):
            import functools
            @functools.wraps(f)
            def decorated(*args, **kwargs):
                if not _is_authenticated():
                    if request.path.startswith("/api/"):
                        return jsonify({"error": "Authentication required"}), 401
                    return redirect("/login")
                perms = _get_user_permissions()
                if permission not in perms and "admin" not in perms:
                    if request.path.startswith("/api/"):
                        return jsonify({"error": f"Access denied. Required permission: '{permission}'"}), 403
                    return redirect("/")
                return f(*args, **kwargs)
            return decorated
        return decorator

    # --- Authentication Routes ---

    @app.route("/login", methods=["GET", "POST"])
    def login():
        """Login page and authentication handler."""
        auth = _get_auth_manager()

        if request.method == "GET":
            if auth and auth.is_authenticated():
                return redirect("/")
            return render_template("login.html", error=None)

        # POST - validate credentials
        username = request.form.get("username", "")
        password = request.form.get("password", "")

        # Try IAM authentication first (multi-user)
        iam = app.config.get("IAM_MANAGER")
        if iam:
            user = iam.authenticate(username, password)
            if user:
                # Create session with role info
                from flask import session as flask_session
                flask_session["authenticated"] = True
                flask_session["username"] = username
                flask_session["role"] = user["role"]
                return redirect("/")

        # Fallback to legacy AuthManager (single-user from .env)
        if auth and auth.authenticate(username, password):
            auth.create_session(username)
            return redirect("/")

        return render_template(
            "login.html", error="Invalid username or password"
        )

    @app.route("/logout", methods=["POST"])
    def logout():
        """Logout and clear session."""
        auth = _get_auth_manager()
        if auth:
            auth.destroy_session()
        return redirect("/login")

    # --- Chat Endpoint ---

    @app.route("/chat", methods=["POST"])
    def chat():
        """Handle chat messages, optionally using a custom assistant.

        Accepts JSON with 'message', optional 'session_id', and optional 'assistant_id'.
        When assistant_id is provided, uses that assistant's system prompt and params.
        """
        data = request.get_json(silent=True)

        # Validate request body
        if data is None or "message" not in data:
            return (
                jsonify({"error": "Missing required field: 'message'"}),
                400,
            )

        message = data["message"]
        session_id = data.get("session_id") or str(uuid.uuid4())
        assistant_id = data.get("assistant_id")

        # Use conversation manager if available
        conversation_manager = app.config.get("CONVERSATION_MANAGER")
        if not conversation_manager:
            return jsonify({
                "response": "The conversation system is being initialized, Sir.",
                "session_id": session_id,
                "agent_steps": [],
            })

        # If a custom assistant is selected, apply its config
        custom_prompt = None
        custom_params = None
        if assistant_id:
            cam = app.config.get("CUSTOM_ASSISTANTS")
            if cam:
                assistant = cam.get_assistant(int(assistant_id))
                if assistant:
                    custom_prompt = assistant["system_prompt"]
                    custom_params = assistant["inference_params"]

        try:
            # Apply custom assistant overrides
            original_params = None
            if custom_params:
                from app.llm_client import inference_params
                original_params = inference_params.to_dict()
                for k, v in custom_params.items():
                    if hasattr(inference_params, k):
                        setattr(inference_params, k, v)

            if custom_prompt:
                conversation_manager._custom_system_prompt = custom_prompt

            response_text = conversation_manager.handle_message(message, session_id)
            agent_steps = conversation_manager._last_agent_steps
            conversation_manager._last_agent_steps = []
        except Exception as e:
            logger.exception("Conversation error while handling session %s", session_id)
            response_text = (
                "I apologize, Sir. An internal error has occurred. "
                "Please try again momentarily."
            )
            agent_steps = []
        finally:
            # Restore original settings
            if custom_prompt:
                conversation_manager._custom_system_prompt = None
            if original_params:
                from app.llm_client import inference_params
                for k, v in original_params.items():
                    if hasattr(inference_params, k):
                        setattr(inference_params, k, v)

        return jsonify({"response": response_text, "session_id": session_id, "agent_steps": agent_steps})

    @app.route("/api/chat-progress/<session_id>", methods=["GET"])
    def chat_progress(session_id):
        """Return gateway lifecycle events collected for a chat session."""
        conversation_manager = app.config.get("CONVERSATION_MANAGER")
        if not conversation_manager:
            return jsonify([])
        return jsonify(conversation_manager.get_progress_events(session_id))

    @app.route("/chat/stream", methods=["POST"])
    def chat_stream():
        """Streaming chat endpoint — returns Server-Sent Events (SSE).

        Each event is a text token: data: <token>
        The final event is:          data: [DONE]
        A session_id event fires first so the client can track it.

        The frontend reads these via fetch + ReadableStream.  TTS is
        triggered client-side only after [DONE] arrives, so audio is
        never split across partial tokens.
        """
        import json as _json
        from app.llm_client import GatewayClient, FailoverLLMClient, inference_params

        data = request.get_json(silent=True)
        if data is None or "message" not in data:
            return jsonify({"error": "Missing 'message'"}), 400

        message = data["message"]
        session_id = data.get("session_id") or str(uuid.uuid4())
        assistant_id = data.get("assistant_id")

        conversation_manager = app.config.get("CONVERSATION_MANAGER")
        if not conversation_manager:
            # Fallback: not ready yet
            def _not_ready():
                yield "data: The conversation system is initialising, Sir.\n\n"
                yield "data: [DONE]\n\n"
            return app.response_class(_not_ready(), mimetype="text/event-stream",
                                       headers={"X-Session-Id": session_id,
                                                "Cache-Control": "no-cache",
                                                "X-Accel-Buffering": "no"})

        # Resolve the GatewayClient (may be wrapped in FailoverLLMClient)
        llm = conversation_manager.llm_client
        gateway: GatewayClient | None = None
        if isinstance(llm, GatewayClient):
            gateway = llm
        elif isinstance(llm, FailoverLLMClient):
            label, first = llm.clients[0]
            if isinstance(first, GatewayClient):
                gateway = first

        if gateway is None:
            # No gateway available — fall back to blocking then fake-stream
            def _blocking_stream():
                response_text = conversation_manager.handle_message(message, session_id)
                # Emit in ~8-word chunks so the UI still animates
                words = response_text.split(" ")
                chunk = []
                for word in words:
                    chunk.append(word)
                    if len(chunk) >= 8:
                        yield f"data: {' '.join(chunk)} \n\n"
                        chunk = []
                if chunk:
                    yield f"data: {' '.join(chunk)}\n\n"
                yield "data: [DONE]\n\n"

            return app.response_class(
                _blocking_stream(), mimetype="text/event-stream",
                headers={"X-Session-Id": session_id,
                         "Cache-Control": "no-cache",
                         "X-Accel-Buffering": "no"})

        # Apply custom assistant overrides if requested
        original_params = None
        if assistant_id:
            cam = app.config.get("CUSTOM_ASSISTANTS")
            if cam:
                assistant = cam.get_assistant(int(assistant_id))
                if assistant:
                    if assistant.get("inference_params"):
                        original_params = inference_params.to_dict()
                        for k, v in assistant["inference_params"].items():
                            if hasattr(inference_params, k):
                                setattr(inference_params, k, v)
                    if assistant.get("system_prompt"):
                        conversation_manager._custom_system_prompt = assistant["system_prompt"]

        def generate():
            try:
                # Send session_id first so the client can store it immediately
                yield f"data: __session__:{session_id}\n\n"

                # Build the message list the same way handle_message does,
                # but drive the LLM call ourselves via stream_chat
                messages = conversation_manager._build_messages(session_id, message)

                # Check think-mode toggles (mirrors GatewayClient.chat logic)
                user_msgs = [m for m in messages if m.get("role") == "user"]
                if user_msgs:
                    last = user_msgs[-1].get("content", "")
                    from app.llm_client import _THINK_ON_RE, _THINK_OFF_RE
                    if _THINK_ON_RE.search(last):
                        gateway._think_mode = True
                    elif _THINK_OFF_RE.search(last):
                        gateway._think_mode = False

                full_text = []
                for token in gateway.stream_chat(messages):
                    full_text.append(token)
                    # Escape newlines inside SSE data field
                    safe = token.replace("\n", "\\n")
                    yield f"data: {safe}\n\n"

                assembled = "".join(full_text)

                # Persist the exchange (history, title generation)
                conversation_manager._persist_exchange(session_id, message, assembled)

            except Exception as exc:
                logger.error("Streaming chat error: %s", exc)
                yield f"data: [ERROR] {exc}\n\n"
            finally:
                # Restore custom assistant settings
                if original_params:
                    for k, v in original_params.items():
                        if hasattr(inference_params, k):
                            setattr(inference_params, k, v)
                if assistant_id:
                    conversation_manager._custom_system_prompt = None

        return app.response_class(
            generate(),
            mimetype="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "X-Accel-Buffering": "no",   # disables nginx buffering
            },
        )

    # --- Health Check ---

    @app.route("/health", methods=["GET"])
    def health():
        """Health check endpoint."""
        from app import _get_capabilities

        config = app.config["JARVIS_CONFIG"]
        capabilities = _get_capabilities(config)

        return jsonify(
            {
                "status": "ok",
                "provider": config.llm_provider,
                "capabilities": capabilities,
            }
        )

    # --- Frontend Routes (Protected) ---

    @app.route("/", methods=["GET"])
    @_login_required
    def index():
        """Chat interface."""
        cfg = app.config.get("CONFIG")
        streaming = cfg.chat_streaming if cfg else False
        return render_template("chat.html", chat_streaming=streaming)

    @app.route("/dashboard", methods=["GET"])
    @_permission_required("monitoring")
    def dashboard():
        """KPI dashboard."""
        return render_template("dashboard.html")

    @app.route("/settings", methods=["GET"])
    @_permission_required("config")
    def settings():
        """Settings page."""
        return render_template("settings.html")

    # --- API Endpoints (Protected) ---

    @app.route("/api/config", methods=["GET"])
    @_login_required
    def get_config():
        """Return current non-sensitive configuration values."""
        config = app.config["JARVIS_CONFIG"]

        return jsonify(
            {
                "provider": config.llm_provider,
                "port": config.port,
                "command_timeout": config.command_timeout,
                "scan_timeout": config.scan_timeout,
                "max_history_pairs": config.max_history_pairs,
                "calendar_provider": config.calendar_provider,
                "reminder_window_minutes": config.reminder_window_minutes,
                "database_path": config.database_path,
                "retention_days": config.retention_days,
                "agent_mode_enabled": config.agent_mode_enabled,
            }
        )

    @app.route("/api/config", methods=["PUT"])
    @_login_required
    def update_config():
        """Update a configuration parameter."""
        data = request.get_json(silent=True)

        if data is None or "key" not in data or "value" not in data:
            return (
                jsonify({"error": "Missing required fields: 'key' and 'value'"}),
                400,
            )

        key = data["key"]
        value = data["value"]
        config = app.config["JARVIS_CONFIG"]

        # Allowed updatable keys and their validation
        updatable_keys = {
            "command_timeout": lambda v: isinstance(v, int) and v > 0,
            "scan_timeout": lambda v: isinstance(v, int) and v > 0,
            "max_history_pairs": lambda v: isinstance(v, int) and v > 0,
            "reminder_window_minutes": lambda v: isinstance(v, int) and v > 0,
            "retention_days": lambda v: isinstance(v, int) and v > 0,
            "agent_mode_enabled": lambda v: isinstance(v, bool),
        }

        if key not in updatable_keys:
            return (
                jsonify({"error": f"Configuration key '{key}' is not updatable"}),
                400,
            )

        if not updatable_keys[key](value):
            expected = "a boolean" if key == "agent_mode_enabled" else "a positive integer"
            return (
                jsonify(
                    {"error": f"Invalid value for '{key}': must be {expected}"}
                ),
                400,
            )

        setattr(config, key, value)
        logger.info("Configuration updated: %s = %s", key, value)

        return jsonify({"success": True, "message": f"Updated '{key}' successfully"})

    @app.route("/api/metrics", methods=["GET"])
    @_login_required
    def get_metrics():
        """Return usage metrics."""
        metrics_collector = app.config.get("METRICS_COLLECTOR")
        if metrics_collector:
            return jsonify(metrics_collector.get_summary())

        return jsonify(
            {
                "total_calls": 0,
                "calls_today": 0,
                "avg_response_ms": 0.0,
                "p95_response_ms": 0.0,
                "tool_usage": {},
                "error_rate": 0.0,
                "active_sessions": 0,
            }
        )

    @app.route("/api/system", methods=["GET"])
    @_login_required
    def get_system():
        """Return system resource usage."""
        system_monitor = app.config.get("SYSTEM_MONITOR")
        if system_monitor:
            metrics = system_monitor.get_metrics()
            return jsonify(
                {
                    "cpu_percent": metrics.cpu_percent,
                    "ram_used_mb": metrics.ram_used_mb,
                    "ram_total_mb": metrics.ram_total_mb,
                    "ram_percent": metrics.ram_percent,
                    "disk_used_gb": metrics.disk_used_gb,
                    "disk_total_gb": metrics.disk_total_gb,
                    "disk_percent": metrics.disk_percent,
                    "cpu_temp_celsius": metrics.cpu_temp_celsius,
                }
            )

        return jsonify(
            {
                "cpu_percent": 0.0,
                "ram_used_mb": 0.0,
                "ram_total_mb": 0.0,
                "ram_percent": 0.0,
                "disk_used_gb": 0.0,
                "disk_total_gb": 0.0,
                "disk_percent": 0.0,
                "cpu_temp_celsius": -1.0,
            }
        )

    @app.route("/api/sessions", methods=["GET"])
    @_login_required
    def get_sessions():
        """Return list of past conversation sessions (filtered by user, admins see all)."""
        db_manager = app.config.get("DB_MANAGER")
        if not db_manager:
            return jsonify([])
        perms = _get_user_permissions()
        username = session.get("username", "")
        # Admins see all sessions, regular users only see their own
        if "admin" in perms:
            sessions = db_manager.get_sessions(limit=50, username=None)
        else:
            sessions = db_manager.get_sessions(limit=50, username=username)
        return jsonify(sessions)

    @app.route("/api/sessions/<session_id>", methods=["GET"])
    @_login_required
    def get_session_history(session_id):
        """Return full message history for a specific session."""
        db_manager = app.config.get("DB_MANAGER")
        if not db_manager:
            return jsonify([])
        history = db_manager.get_full_history(session_id)
        return jsonify(history)

    @app.route("/api/notifications", methods=["GET"])
    @_login_required
    def get_notifications():
        """Return pending reminder notifications from the scheduler."""
        scheduler = app.config.get("SCHEDULER")
        if not scheduler:
            return jsonify([])
        notifications = scheduler.get_pending_notifications()
        return jsonify(notifications)

    @app.route("/api/notes/<int:note_id>/snooze", methods=["POST"])
    @_login_required
    def snooze_note(note_id):
        """Snooze a reminder: acknowledge it now and reschedule for later.

        Body: {"minutes": 5}  (how many minutes to postpone)
        """
        from datetime import datetime, timedelta
        import sqlite3

        data = request.get_json(silent=True)
        if not data or "minutes" not in data:
            return jsonify({"error": "Missing 'minutes' field"}), 400

        minutes = int(data["minutes"])
        if minutes < 1 or minutes > 1440:
            return jsonify({"error": "Minutes must be between 1 and 1440"}), 400

        scheduler = app.config.get("SCHEDULER")
        db_manager = app.config.get("DB_MANAGER")
        if not scheduler or not db_manager:
            return jsonify({"error": "Scheduler not available"}), 500

        new_due = datetime.now() + timedelta(minutes=minutes)
        new_due_str = new_due.strftime("%Y-%m-%d %H:%M:%S")

        try:
            conn = sqlite3.connect(db_manager.db_path)
            conn.row_factory = sqlite3.Row

            # Verify note exists
            row = conn.execute("SELECT * FROM notes WHERE id = ?", (note_id,)).fetchone()
            if not row:
                conn.close()
                return jsonify({"error": "Note not found"}), 404

            # Reschedule the due_date
            conn.execute("UPDATE notes SET due_date = ?, status = 'active' WHERE id = ?",
                         (new_due_str, note_id))

            # Extend expires_at if it exists and is set
            if row["expires_at"] if "expires_at" in row.keys() else None:
                new_expires = new_due + timedelta(minutes=1)
                conn.execute("UPDATE notes SET expires_at = ? WHERE id = ?",
                             (new_expires.strftime("%Y-%m-%d %H:%M:%S"), note_id))

            conn.commit()
            conn.close()

            # Remove from fired_ids so it can fire again
            scheduler._fired_note_ids.discard(note_id)

            # Remove from unacknowledged (it's been snoozed, not pending anymore)
            with scheduler._lock:
                if note_id in scheduler.unacknowledged:
                    del scheduler.unacknowledged[note_id]

            logger.info("Note #%d snoozed for %d minutes (new due: %s)", note_id, minutes, new_due_str)
            return jsonify({
                "success": True,
                "new_due": new_due.strftime("%H:%M"),
                "message": f"Snoozed for {minutes} minute{'s' if minutes != 1 else ''}."
            })

        except Exception as e:
            logger.error("Snooze failed for note #%d: %s", note_id, e)
            return jsonify({"error": str(e)}), 500

    @app.route("/api/upload", methods=["POST"])
    @_login_required
    def upload_file():
        """Upload a file and extract its text content for use as context.
        Also saves the file to uploads/ so it can be referenced later."""
        from app.context_extractor import extract_from_file
        import os

        if "file" not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files["file"]
        if not file.filename:
            return jsonify({"error": "Empty filename"}), 400

        # Save the file to uploads/ folder for future reference
        project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        uploads_dir = os.path.join(project_dir, "uploads")
        os.makedirs(uploads_dir, exist_ok=True)
        save_path = os.path.join(uploads_dir, file.filename)
        file.save(save_path)
        file.seek(0)  # Reset for extraction

        result = extract_from_file(file)
        if "error" in result:
            return jsonify(result), 400

        result["saved_path"] = save_path
        return jsonify(result)

    @app.route("/api/fetch-url", methods=["POST"])
    @_login_required
    def fetch_url():
        """Fetch a URL and extract its text content for use as context."""
        from app.context_extractor import extract_from_url

        data = request.get_json(silent=True)
        if not data or "url" not in data:
            return jsonify({"error": "Missing 'url' field"}), 400

        result = extract_from_url(data["url"])
        if "error" in result:
            return jsonify(result), 400

        return jsonify(result)

    @app.route("/chat-with-context", methods=["POST"])
    def chat_with_context():
        """Handle chat messages with additional context attachments.

        Accepts JSON with 'message', optional 'session_id', and optional 'context' array.
        Each context item has 'type' (file/url), 'name', and 'content'.
        """
        data = request.get_json(silent=True)

        if data is None or "message" not in data:
            return jsonify({"error": "Missing required field: 'message'"}), 400

        message = data["message"]
        session_id = data.get("session_id") or str(uuid.uuid4())
        context_items = data.get("context", [])

        # Build context prefix from attachments
        context_prefix = ""
        if context_items:
            parts = ["[ATTACHED CONTEXT — use this information to answer the user's question]\n"]
            for item in context_items:
                source = item.get("name", "unknown")
                content = item.get("content", "")
                parts.append(f"--- {source} ---\n{content}\n")
            context_prefix = "\n".join(parts) + "\n[END OF ATTACHED CONTEXT]\n\n"

        # Prepend context to the message for the conversation manager
        full_message = context_prefix + message if context_prefix else message

        conversation_manager = app.config.get("CONVERSATION_MANAGER")
        if conversation_manager:
            try:
                response_text = conversation_manager.handle_message(full_message, session_id)
            except Exception as e:
                logger.error("Conversation error: %s", e)
                response_text = "I apologize, Sir. An internal error has occurred."
        else:
            response_text = "The conversation system is being initialized, Sir."

        return jsonify({"response": response_text, "session_id": session_id})

    @app.route("/api/plugins", methods=["GET"])
    @_login_required
    def list_plugins():
        """List all loaded plugins."""
        pm = app.config.get("PLUGIN_MANAGER")
        if not pm:
            return jsonify([])
        return jsonify(pm.list_plugins())

    @app.route("/api/plugins", methods=["POST"])
    @_login_required
    def install_plugin():
        """Install a new plugin from submitted Python code."""
        pm = app.config.get("PLUGIN_MANAGER")
        if not pm:
            return jsonify({"error": "Plugin system not available"}), 500

        data = request.get_json(silent=True)
        if not data or "name" not in data or "code" not in data:
            return jsonify({"error": "Missing 'name' and 'code' fields"}), 400

        result = pm.install_plugin(data["name"], data["code"])
        if "error" in result:
            return jsonify(result), 400
        return jsonify(result)

    @app.route("/api/plugins/<name>", methods=["DELETE"])
    @_login_required
    def uninstall_plugin(name):
        """Uninstall a plugin by name."""
        pm = app.config.get("PLUGIN_MANAGER")
        if not pm:
            return jsonify({"error": "Plugin system not available"}), 500
        result = pm.uninstall_plugin(name)
        if "error" in result:
            return jsonify(result), 404
        return jsonify(result)

    @app.route("/api/inference", methods=["GET"])
    @_login_required
    def get_inference_params():
        """Return current inference parameters."""
        from app.llm_client import inference_params
        return jsonify(inference_params.to_dict())

    @app.route("/api/inference", methods=["PUT"])
    @_login_required
    def update_inference_params():
        """Update inference parameters at runtime."""
        from app.llm_client import inference_params

        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "No data provided"}), 400

        # Validate and apply each parameter
        errors = []
        if "temperature" in data:
            val = data["temperature"]
            if isinstance(val, (int, float)) and 0.0 <= val <= 2.0:
                inference_params.temperature = float(val)
            else:
                errors.append("temperature must be between 0.0 and 2.0")

        if "top_p" in data:
            val = data["top_p"]
            if isinstance(val, (int, float)) and 0.0 <= val <= 1.0:
                inference_params.top_p = float(val)
            else:
                errors.append("top_p must be between 0.0 and 1.0")

        if "max_tokens" in data:
            val = data["max_tokens"]
            if isinstance(val, int) and 1 <= val <= 4096:
                inference_params.max_tokens = val
            else:
                errors.append("max_tokens must be between 1 and 4096")

        if "frequency_penalty" in data:
            val = data["frequency_penalty"]
            if isinstance(val, (int, float)) and -2.0 <= val <= 2.0:
                inference_params.frequency_penalty = float(val)
            else:
                errors.append("frequency_penalty must be between -2.0 and 2.0")

        if "presence_penalty" in data:
            val = data["presence_penalty"]
            if isinstance(val, (int, float)) and -2.0 <= val <= 2.0:
                inference_params.presence_penalty = float(val)
            else:
                errors.append("presence_penalty must be between -2.0 and 2.0")

        if errors:
            return jsonify({"error": "; ".join(errors)}), 400

        logger.info("Inference params updated: %s", inference_params.to_dict())
        return jsonify({"success": True, "params": inference_params.to_dict()})

    @app.route("/api/search", methods=["GET"])
    @_login_required
    def search_conversations():
        """Search across all past conversations."""
        query = request.args.get("q", "").strip()
        if not query:
            return jsonify([])
        db_manager = app.config.get("DB_MANAGER")
        if not db_manager:
            return jsonify([])
        results = db_manager.search_conversations(query, limit=30)
        return jsonify(results)

    @app.route("/api/metrics/daily", methods=["GET"])
    @_login_required
    def get_daily_metrics():
        """Return per-day metrics breakdown for the last 7 days."""
        metrics_collector = app.config.get("METRICS_COLLECTOR")
        if metrics_collector:
            return jsonify(metrics_collector.get_daily_breakdown(days=7))
        return jsonify([])

    @app.route("/api/metrics/hourly", methods=["GET"])
    @_login_required
    def get_hourly_metrics():
        """Return per-hour metrics breakdown for the last 24 hours."""
        db_manager = app.config.get("DB_MANAGER")
        if db_manager:
            return jsonify(db_manager.get_hourly_breakdown(hours=24))
        return jsonify([])

    @app.route("/api/sessions/<session_id>/export", methods=["GET"])
    @_login_required
    def export_session_xlsx(session_id):
        """Export a conversation session as an XLSX file."""
        import io
        from flask import send_file

        db_manager = app.config.get("DB_MANAGER")
        if not db_manager:
            return jsonify({"error": "Database not available"}), 500

        history = db_manager.get_full_history(session_id)
        if not history:
            return jsonify({"error": "Session not found or empty"}), 404

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Conversation"

            # Header row
            headers = ["Timestamp", "Role", "Message"]
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
            header_font_white = Font(bold=True, color="FFFFFF")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row_idx, msg in enumerate(history, 2):
                ws.cell(row=row_idx, column=1, value=msg.get("timestamp", ""))
                ws.cell(row=row_idx, column=2, value=msg.get("role", "").capitalize())
                cell = ws.cell(row=row_idx, column=3, value=msg.get("content", ""))
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Column widths
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 80

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename = f"jarvis_conversation_{session_id[:8]}.xlsx"
            return send_file(
                buffer,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename,
            )

        except ImportError:
            # Fallback: export as CSV if openpyxl not installed
            import csv

            buffer = io.StringIO()
            writer = csv.writer(buffer)
            writer.writerow(["Timestamp", "Role", "Message"])
            for msg in history:
                writer.writerow([msg.get("timestamp", ""), msg.get("role", ""), msg.get("content", "")])

            csv_buffer = io.BytesIO(buffer.getvalue().encode("utf-8"))
            csv_buffer.seek(0)
            filename = f"jarvis_conversation_{session_id[:8]}.csv"
            return send_file(
                csv_buffer,
                mimetype="text/csv",
                as_attachment=True,
                download_name=filename,
            )

    @app.route("/api/docs", methods=["GET"])
    def swagger_ui():
        """Serve the Swagger UI documentation page."""
        from app.swagger import SWAGGER_UI_HTML
        return SWAGGER_UI_HTML, 200, {"Content-Type": "text/html"}

    @app.route("/api/spec", methods=["GET"])
    def swagger_spec():
        """Serve the OpenAPI JSON specification (includes dynamic script endpoints)."""
        from app.swagger import OPENAPI_SPEC
        import copy
        spec = copy.deepcopy(OPENAPI_SPEC)

        # Merge dynamic script endpoints
        script_runner = app.config.get("SCRIPT_RUNNER")
        if script_runner:
            script_paths = script_runner.get_swagger_paths()
            spec["paths"].update(script_paths)

        return jsonify(spec)

    @app.route("/api/scripts", methods=["GET"])
    @_login_required
    def list_scripts():
        """List all available scripts in the scripts/ folder."""
        script_runner = app.config.get("SCRIPT_RUNNER")
        if not script_runner:
            return jsonify([])
        return jsonify(script_runner.list_scripts())

    @app.route("/api/scripts/<script_name>", methods=["POST"])
    @_login_required
    def run_script(script_name):
        """Execute a script from the scripts/ folder by name."""
        script_runner = app.config.get("SCRIPT_RUNNER")
        if not script_runner:
            return jsonify({"error": "Script runner not available"}), 500

        args = request.get_json(silent=True) or {}
        result = script_runner.run_script(script_name, args=args)

        if "error" in result:
            return jsonify(result), 404 if "not found" in result["error"].lower() else 500

        return jsonify(result)

    @app.route("/api/scripts/reload", methods=["POST"])
    @_login_required
    def reload_scripts():
        """Re-scan the scripts/ folder for new or removed scripts."""
        script_runner = app.config.get("SCRIPT_RUNNER")
        if not script_runner:
            return jsonify({"error": "Script runner not available"}), 500
        count = script_runner.discover_scripts()
        return jsonify({"message": f"Discovered {count} script(s)", "scripts": script_runner.list_scripts()})

    @app.route("/api/subscribe", methods=["GET"])
    def subscribe_sse():
        """Server-Sent Events endpoint for real-time notifications.

        Clients connect and receive events when:
        - A reminder fires
        - A scheduled command completes
        - The assistant sends a proactive message

        Usage: const evtSource = new EventSource('/api/subscribe');
               evtSource.onmessage = (e) => console.log(JSON.parse(e.data));
        """
        from flask import Response
        import time as _time
        import json as _json

        def event_stream():
            """Generator that yields SSE events."""
            scheduler = app.config.get("SCHEDULER")
            last_check = _time.time()

            # Send initial connection event
            yield f"data: {_json.dumps({'type': 'connected', 'message': 'Subscribed to Jarvis notifications'})}\n\n"

            while True:
                _time.sleep(2)  # Check every 2 seconds

                if scheduler:
                    notifications = scheduler.get_pending_notifications()
                    for notif in notifications:
                        yield f"data: {_json.dumps(notif)}\n\n"

                # Send heartbeat every 30 seconds to keep connection alive
                if _time.time() - last_check > 30:
                    yield f": heartbeat\n\n"
                    last_check = _time.time()

        return Response(
            event_stream(),
            mimetype="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no",
            },
        )

    # ── Database Management API ────────────────────────────────────────────

    @app.route("/api/db/tables", methods=["GET"])
    @_login_required
    def db_list_tables():
        """List all tables in the database with row counts."""
        db_manager = app.config.get("DB_MANAGER")
        if not db_manager:
            return jsonify({"error": "Database not available"}), 500
        conn = db_manager._get_connection()
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = []
        for row in cursor.fetchall():
            name = row["name"]
            count = conn.execute(f"SELECT COUNT(*) as cnt FROM [{name}]").fetchone()["cnt"]
            tables.append({"name": name, "row_count": count})
        return jsonify(tables)

    @app.route("/api/db/<table_name>", methods=["GET"])
    @_login_required
    def db_get_rows(table_name):
        """Get all rows from a table (with optional limit/offset)."""
        db_manager = app.config.get("DB_MANAGER")
        if not db_manager:
            return jsonify({"error": "Database not available"}), 500

        # Validate table name (prevent SQL injection)
        conn = db_manager._get_connection()
        valid_tables = [r["name"] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        if table_name not in valid_tables:
            return jsonify({"error": f"Table '{table_name}' not found"}), 404

        limit = request.args.get("limit", "100", type=str)
        offset = request.args.get("offset", "0", type=str)

        # Get column info
        columns = [r[1] for r in conn.execute(f"PRAGMA table_info([{table_name}])").fetchall()]

        cursor = conn.execute(f"SELECT * FROM [{table_name}] ORDER BY rowid DESC LIMIT ? OFFSET ?", (int(limit), int(offset)))
        rows = [dict(row) for row in cursor.fetchall()]

        total = conn.execute(f"SELECT COUNT(*) as cnt FROM [{table_name}]").fetchone()["cnt"]

        return jsonify({"table": table_name, "columns": columns, "rows": rows, "total": total})

    @app.route("/api/db/<table_name>", methods=["POST"])
    @_login_required
    def db_insert_row(table_name):
        """Insert a new row into a table."""
        db_manager = app.config.get("DB_MANAGER")
        if not db_manager:
            return jsonify({"error": "Database not available"}), 500

        conn = db_manager._get_connection()
        valid_tables = [r["name"] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        if table_name not in valid_tables:
            return jsonify({"error": f"Table '{table_name}' not found"}), 404

        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "No data provided"}), 400

        columns = list(data.keys())
        values = list(data.values())
        placeholders = ", ".join(["?"] * len(columns))
        col_names = ", ".join([f"[{c}]" for c in columns])

        try:
            cursor = conn.execute(f"INSERT INTO [{table_name}] ({col_names}) VALUES ({placeholders})", values)
            conn.commit()
            return jsonify({"success": True, "id": cursor.lastrowid, "message": f"Row inserted into {table_name}"})
        except Exception as e:
            return jsonify({"error": f"Insert failed: {e}"}), 400

    @app.route("/api/db/<table_name>/<int:row_id>", methods=["PUT"])
    @_login_required
    def db_update_row(table_name, row_id):
        """Update a row by its rowid/id."""
        db_manager = app.config.get("DB_MANAGER")
        if not db_manager:
            return jsonify({"error": "Database not available"}), 500

        conn = db_manager._get_connection()
        valid_tables = [r["name"] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        if table_name not in valid_tables:
            return jsonify({"error": f"Table '{table_name}' not found"}), 404

        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "No data provided"}), 400

        set_clauses = ", ".join([f"[{k}] = ?" for k in data.keys()])
        values = list(data.values()) + [row_id]

        try:
            cursor = conn.execute(f"UPDATE [{table_name}] SET {set_clauses} WHERE id = ?", values)
            conn.commit()
            if cursor.rowcount == 0:
                return jsonify({"error": f"Row {row_id} not found in {table_name}"}), 404
            return jsonify({"success": True, "message": f"Row {row_id} updated in {table_name}"})
        except Exception as e:
            return jsonify({"error": f"Update failed: {e}"}), 400

    @app.route("/api/db/<table_name>/<int:row_id>", methods=["DELETE"])
    @_login_required
    def db_delete_row(table_name, row_id):
        """Delete a row by its id."""
        db_manager = app.config.get("DB_MANAGER")
        if not db_manager:
            return jsonify({"error": "Database not available"}), 500

        conn = db_manager._get_connection()
        valid_tables = [r["name"] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        if table_name not in valid_tables:
            return jsonify({"error": f"Table '{table_name}' not found"}), 404

        try:
            cursor = conn.execute(f"DELETE FROM [{table_name}] WHERE id = ?", (row_id,))
            conn.commit()
            if cursor.rowcount == 0:
                return jsonify({"error": f"Row {row_id} not found in {table_name}"}), 404
            return jsonify({"success": True, "message": f"Row {row_id} deleted from {table_name}"})
        except Exception as e:
            return jsonify({"error": f"Delete failed: {e}"}), 400

    @app.route("/database", methods=["GET"])
    @_login_required
    def database_page():
        """Database management UI page."""
        return render_template("database.html")

    # ── Server Restart API ─────────────────────────────────────────────────

    @app.route("/api/restart/status", methods=["GET"])
    def restart_status():
        """Check if a restart is pending (used by frontend for snackbar)."""
        file_watcher = app.config.get("FILE_WATCHER")
        if file_watcher and file_watcher._restart_pending:
            import time as _t
            remaining = max(0, file_watcher._restart_time - _t.time())
            return jsonify({"pending": True, "remaining_seconds": round(remaining, 1)})
        return jsonify({"pending": False, "remaining_seconds": 0})

    @app.route("/api/restart", methods=["POST"])
    @_login_required
    def manual_restart():
        """Manually trigger a server restart."""
        import threading
        import time as _t

        def do_restart():
            _t.sleep(3)
            logger.info("Manual restart triggered")
            file_watcher = app.config.get("FILE_WATCHER")
            if file_watcher:
                file_watcher._restart()
            else:
                import os, sys
                os._exit(0)

        # Set pending flag
        file_watcher = app.config.get("FILE_WATCHER")
        if file_watcher:
            file_watcher._restart_pending = True
            file_watcher._restart_time = _t.time() + 3

        threading.Thread(target=do_restart, daemon=True).start()
        return jsonify({"success": True, "message": "Server restarting in 3 seconds..."})

    # ── Device Permissions & Location API ──────────────────────────────────

    @app.route("/api/device/location", methods=["POST"])
    def report_location():
        """Report device GPS location to the server.

        Called by the frontend when GPS permission is granted.
        Stores the latest known position for each device.
        """
        data = request.get_json(silent=True)
        if not data or "latitude" not in data or "longitude" not in data:
            return jsonify({"error": "Missing latitude/longitude"}), 400

        device_id = data.get("device_id", request.remote_addr)
        locations = app.config.setdefault("DEVICE_LOCATIONS", {})
        locations[device_id] = {
            "latitude": data["latitude"],
            "longitude": data["longitude"],
            "accuracy": data.get("accuracy"),
            "timestamp": data.get("timestamp", ""),
            "device_id": device_id,
        }

        # Evaluate GPS-based workflows
        engine = app.config.get("WORKFLOW_ENGINE")
        if engine:
            try:
                engine.evaluate_gps(device_id, data["latitude"], data["longitude"])
            except Exception as e:
                logger.warning("GPS workflow evaluation error: %s", e)

        return jsonify({"success": True})

    @app.route("/api/device/locations", methods=["GET"])
    @_login_required
    def get_device_locations():
        """Get all known device locations."""
        locations = app.config.get("DEVICE_LOCATIONS", {})
        return jsonify(list(locations.values()))

    # ── Bluetooth API ──────────────────────────────────────────────────────

    @app.route("/api/bluetooth/scan", methods=["POST"])
    @_login_required
    def bt_scan():
        """Scan for nearby Bluetooth devices."""
        bt = app.config.get("BLUETOOTH_MANAGER")
        if not bt:
            return jsonify({"error": "Bluetooth manager not available"}), 500
        duration = request.get_json(silent=True) or {}
        devices = bt.scan(duration=duration.get("duration", 5))
        return jsonify({"devices": devices, "scanning": False})

    @app.route("/api/bluetooth/devices", methods=["GET"])
    @_login_required
    def bt_devices():
        """List all known Bluetooth devices."""
        bt = app.config.get("BLUETOOTH_MANAGER")
        if not bt:
            return jsonify([])
        return jsonify(bt.get_devices())

    @app.route("/api/bluetooth/sensor-data", methods=["POST"])
    def bt_report_sensor():
        """Report IMU sensor data from a connected Bluetooth device.

        Called by scripts reading BLE IMU data to feed it into Jarvis.
        """
        bt = app.config.get("BLUETOOTH_MANAGER")
        if not bt:
            return jsonify({"error": "Bluetooth manager not available"}), 500

        data = request.get_json(silent=True)
        if not data or "address" not in data:
            return jsonify({"error": "Missing 'address' field"}), 400

        bt.report_sensor_data(data["address"], data)
        return jsonify({"success": True})

    @app.route("/api/bluetooth/sensor-data/<address>", methods=["GET"])
    @_login_required
    def bt_get_sensor_data(address):
        """Get recent sensor readings for a specific device."""
        bt = app.config.get("BLUETOOTH_MANAGER")
        if not bt:
            return jsonify([])
        last_n = request.args.get("last", "10", type=str)
        return jsonify(bt.get_sensor_data(address, last_n=int(last_n)))

    @app.route("/api/bluetooth/context", methods=["GET"])
    def bt_fused_context():
        """Get fused sensor + location context.

        Combines IMU data with phone GPS for room-level positioning.
        Scripts can poll this endpoint to augment their data.
        """
        bt = app.config.get("BLUETOOTH_MANAGER")
        locations = app.config.get("DEVICE_LOCATIONS", {})

        context = bt.get_fused_context() if bt else {"imu_devices": {}, "connected_count": 0}
        context["device_locations"] = list(locations.values())
        return jsonify(context)

    @app.route("/api/bluetooth/connect", methods=["POST"])
    @_login_required
    def bt_connect():
        """Attempt to connect to a Bluetooth device."""
        bt = app.config.get("BLUETOOTH_MANAGER")
        if not bt:
            return jsonify({"error": "Bluetooth manager not available"}), 500
        data = request.get_json(silent=True) or {}
        address = data.get("address", "")
        if not address:
            return jsonify({"error": "Missing 'address' field"}), 400

        # Mark as connected (actual BLE connection would require async bleak connect)
        if address in bt.devices:
            bt.devices[address].connected = True
            bt.connected_devices.add(address)
            return jsonify({"success": True, "message": f"Connected to {address}"})

        return jsonify({"error": "Device not found. Scan first."}), 404

    @app.route("/api/bluetooth/disconnect", methods=["POST"])
    @_login_required
    def bt_disconnect():
        """Disconnect from a Bluetooth device."""
        bt = app.config.get("BLUETOOTH_MANAGER")
        if not bt:
            return jsonify({"error": "Bluetooth manager not available"}), 500
        data = request.get_json(silent=True) or {}
        address = data.get("address", "")
        if not address:
            return jsonify({"error": "Missing 'address' field"}), 400

        bt.connected_devices.discard(address)
        if address in bt.devices:
            bt.devices[address].connected = False
        return jsonify({"success": True, "message": f"Disconnected from {address}"})

    @app.route("/bluetooth", methods=["GET"])
    @_login_required
    def bluetooth_page():
        """Bluetooth management UI page."""
        return render_template("bluetooth.html")

    # ── IAM (Identity & Access Management) API ─────────────────────────────

    def _require_admin(f):
        """Decorator: requires the logged-in user to have admin permission."""
        import functools
        @functools.wraps(f)
        def decorated(*args, **kwargs):
            if not _is_authenticated():
                if request.path.startswith("/api/"):
                    return jsonify({"error": "Authentication required"}), 401
                return redirect("/login")
            perms = _get_user_permissions()
            if "admin" not in perms:
                if request.path.startswith("/api/"):
                    return jsonify({"error": "Admin access required"}), 403
                return redirect("/")
            return f(*args, **kwargs)
        return decorated

    @app.route("/api/iam/users", methods=["GET"])
    @_require_admin
    def iam_list_users():
        """List all users."""
        iam = app.config.get("IAM_MANAGER")
        if not iam:
            return jsonify([])
        return jsonify(iam.list_users())

    @app.route("/api/iam/users", methods=["POST"])
    @_require_admin
    def iam_create_user():
        """Create a new user."""
        iam = app.config.get("IAM_MANAGER")
        if not iam:
            return jsonify({"error": "IAM not available"}), 500
        data = request.get_json(silent=True)
        if not data or "username" not in data or "password" not in data:
            return jsonify({"error": "Missing username or password"}), 400
        result = iam.create_user(data["username"], data["password"], data.get("role", "user"))
        return jsonify(result), 200 if "success" in result else 400

    @app.route("/api/iam/users/<username>", methods=["PUT"])
    @_require_admin
    def iam_update_user(username):
        """Update a user's role, status, or password."""
        iam = app.config.get("IAM_MANAGER")
        if not iam:
            return jsonify({"error": "IAM not available"}), 500
        data = request.get_json(silent=True) or {}
        result = iam.update_user(username, data)
        return jsonify(result), 200 if "success" in result else 400

    @app.route("/api/iam/users/<username>", methods=["DELETE"])
    @_require_admin
    def iam_delete_user(username):
        """Delete a user."""
        iam = app.config.get("IAM_MANAGER")
        if not iam:
            return jsonify({"error": "IAM not available"}), 500
        result = iam.delete_user(username)
        return jsonify(result), 200 if "success" in result else 404

    @app.route("/api/iam/roles", methods=["GET"])
    @_require_admin
    def iam_list_roles():
        """List all roles with permissions."""
        iam = app.config.get("IAM_MANAGER")
        if not iam:
            return jsonify([])
        return jsonify(iam.list_roles())

    @app.route("/api/iam/roles", methods=["POST"])
    @_require_admin
    def iam_update_role():
        """Create or update a role."""
        iam = app.config.get("IAM_MANAGER")
        if not iam:
            return jsonify({"error": "IAM not available"}), 500
        data = request.get_json(silent=True)
        if not data or "name" not in data or "permissions" not in data:
            return jsonify({"error": "Missing name or permissions"}), 400
        result = iam.update_role(data["name"], data["permissions"], data.get("description", ""))
        return jsonify(result)

    @app.route("/api/iam/roles/<name>", methods=["DELETE"])
    @_require_admin
    def iam_delete_role(name):
        """Delete a custom role."""
        iam = app.config.get("IAM_MANAGER")
        if not iam:
            return jsonify({"error": "IAM not available"}), 500
        result = iam.delete_role(name)
        return jsonify(result), 200 if "success" in result else 400

    @app.route("/api/iam/permissions", methods=["GET"])
    @_require_admin
    def iam_list_permissions():
        """List all available permission groups."""
        from app.iam import PERMISSION_GROUPS
        return jsonify(PERMISSION_GROUPS)

    @app.route("/iam", methods=["GET"])
    @_require_admin
    def iam_page():
        """IAM management UI page (admin only)."""
        return render_template("iam.html")

    # ── User Preferences & Voice Enrollment API ────────────────────────────

    @app.route("/api/user/preferences", methods=["GET"])
    @_login_required
    def get_user_preferences():
        """Get the current user's preferences."""
        prefs_mgr = app.config.get("USER_PREFS_MANAGER")
        if not prefs_mgr:
            return jsonify({})
        username = _get_current_username()
        return jsonify(prefs_mgr.get_preferences(username))

    @app.route("/api/user/preferences", methods=["PUT"])
    @_login_required
    def update_user_preferences():
        """Update the current user's preferences."""
        prefs_mgr = app.config.get("USER_PREFS_MANAGER")
        if not prefs_mgr:
            return jsonify({"error": "Preferences not available"}), 500
        username = _get_current_username()
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "No data provided"}), 400
        prefs_mgr.set_preferences(username, data)
        return jsonify({"success": True, "message": "Preferences updated"})

    @app.route("/api/user/voice-enroll", methods=["POST"])
    @_login_required
    def voice_enroll():
        """Submit a voice sample for enrollment.

        The frontend sends audio features (pitch, energy, MFCCs) extracted
        from the user speaking specific phrases. After 3+ samples, the profile
        is considered enrolled.
        """
        prefs_mgr = app.config.get("USER_PREFS_MANAGER")
        if not prefs_mgr:
            return jsonify({"error": "Preferences not available"}), 500

        username = _get_current_username()
        data = request.get_json(silent=True)
        if not data or "features" not in data:
            return jsonify({"error": "Missing 'features' field"}), 400

        # Get existing profile or start new
        existing = prefs_mgr.get_voice_profile(username)
        if existing:
            profile = existing["profile"]
            sample_count = existing["sample_count"] + 1
            # Average the features with existing profile
            for key in data["features"]:
                if key in profile:
                    profile[key] = (profile[key] * (sample_count - 1) + data["features"][key]) / sample_count
                else:
                    profile[key] = data["features"][key]
        else:
            profile = data["features"]
            sample_count = 1

        prefs_mgr.save_voice_profile(username, profile, sample_count)

        enrolled = sample_count >= 3
        return jsonify({
            "success": True,
            "sample_count": sample_count,
            "enrolled": enrolled,
            "message": f"Sample {sample_count} recorded." + (" Voice profile complete!" if enrolled else f" Need {3 - sample_count} more sample(s).")
        })

    @app.route("/api/user/voice-identify", methods=["POST"])
    def voice_identify():
        """Identify a speaker from voice features.

        Compares submitted features against all enrolled profiles.
        Returns the best matching username or null if no match.
        """
        prefs_mgr = app.config.get("USER_PREFS_MANAGER")
        if not prefs_mgr:
            return jsonify({"error": "Preferences not available"}), 500

        data = request.get_json(silent=True)
        if not data or "features" not in data:
            return jsonify({"error": "Missing 'features' field"}), 400

        features = data["features"]
        profiles = prefs_mgr.get_all_voice_profiles()

        if not profiles:
            return jsonify({"match": None, "confidence": 0, "message": "No voice profiles enrolled"})

        # Simple cosine-similarity-based matching
        best_match = None
        best_score = -1

        for username, profile in profiles.items():
            score = _compute_voice_similarity(features, profile)
            if score > best_score:
                best_score = score
                best_match = username

        # Threshold: require at least 0.7 similarity
        if best_score >= 0.7:
            return jsonify({"match": best_match, "confidence": round(best_score, 3)})
        return jsonify({"match": None, "confidence": round(best_score, 3), "message": "No confident match"})

    @app.route("/signup", methods=["GET", "POST"])
    def signup():
        """Self-registration page."""
        if request.method == "GET":
            return render_template("signup.html", error=None)

        # POST - create account
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        honorific = request.form.get("honorific", "Sir").strip()

        if not username or not password:
            return render_template("signup.html", error="Username and password are required")
        if len(password) < 4:
            return render_template("signup.html", error="Password must be at least 4 characters")

        iam = app.config.get("IAM_MANAGER")
        if not iam:
            return render_template("signup.html", error="Registration not available")

        result = iam.create_user(username, password, "user")
        if "error" in result:
            return render_template("signup.html", error=result["error"])

        # Set initial preferences
        prefs_mgr = app.config.get("USER_PREFS_MANAGER")
        if prefs_mgr:
            prefs_mgr.set_preference(username, "honorific", honorific)

        # Auto-login after signup
        from flask import session as flask_session
        flask_session["authenticated"] = True
        flask_session["username"] = username
        flask_session["role"] = "user"
        return redirect("/")

    def _compute_voice_similarity(features_a, features_b):
        """Compute similarity between two voice feature sets (0-1 scale)."""
        common_keys = set(features_a.keys()) & set(features_b.keys())
        if not common_keys:
            return 0.0
        total_diff = 0.0
        for key in common_keys:
            try:
                a = float(features_a[key])
                b = float(features_b[key])
                diff = abs(a - b) / (max(abs(a), abs(b), 1.0))
                total_diff += diff
            except (ValueError, TypeError):
                continue
        avg_diff = total_diff / len(common_keys) if common_keys else 1.0
        return max(0.0, 1.0 - avg_diff)

    # ── Face Authentication API ────────────────────────────────────────────

    @app.route("/api/user/face-enroll", methods=["POST"])
    @_login_required
    def face_enroll():
        """Submit a face descriptor for enrollment.

        The frontend uses face-api.js to detect the face and extract a
        128-dimensional descriptor. The user must submit descriptors for
        5 angles: front, up, down, left, right.
        """
        face_mgr = app.config.get("FACE_PROFILE_MANAGER")
        if not face_mgr:
            return jsonify({"error": "Face authentication not available"}), 500

        data = request.get_json(silent=True)
        if not data or "angle" not in data or "descriptor" not in data:
            return jsonify({"error": "Missing 'angle' and 'descriptor' fields"}), 400

        username = _get_current_username()
        angle = data["angle"]
        descriptor = data["descriptor"]

        valid_angles = {"front", "up", "down", "left", "right"}
        if angle not in valid_angles:
            return jsonify({"error": f"Invalid angle. Must be one of: {valid_angles}"}), 400

        if not isinstance(descriptor, list) or len(descriptor) < 64:
            return jsonify({"error": "Invalid descriptor. Expected a list of 64+ floats."}), 400

        face_mgr.save_face_descriptor(username, angle, descriptor)

        # Check enrollment completeness
        profile = face_mgr.get_face_profile(username)
        enrolled_angles = set(profile.keys())
        remaining = valid_angles - enrolled_angles
        complete = len(remaining) == 0

        return jsonify({
            "success": True,
            "angle": angle,
            "enrolled_angles": list(enrolled_angles),
            "remaining_angles": list(remaining),
            "complete": complete,
            "message": "Face enrollment complete! You can now use face login." if complete else f"Recorded '{angle}'. Still need: {', '.join(remaining)}"
        })

    @app.route("/api/user/face-identify", methods=["POST"])
    def face_identify():
        """Identify a user by their face descriptor.

        Returns the matching username if found, or null.
        Can be used for auto-login via face recognition.
        """
        face_mgr = app.config.get("FACE_PROFILE_MANAGER")
        if not face_mgr:
            return jsonify({"error": "Face authentication not available"}), 500

        data = request.get_json(silent=True)
        if not data or "descriptor" not in data:
            return jsonify({"error": "Missing 'descriptor' field"}), 400

        descriptor = data["descriptor"]
        result = face_mgr.identify_face(descriptor, threshold=0.6)

        if result:
            username, distance = result
            # Auto-login the identified user
            from flask import session as flask_session
            iam = app.config.get("IAM_MANAGER")
            if iam and username in iam.users:
                user = iam.users[username]
                flask_session["authenticated"] = True
                flask_session["username"] = username
                flask_session["role"] = user["role"]

            return jsonify({
                "match": username,
                "distance": round(distance, 4),
                "confidence": round(max(0, 1 - distance) * 100, 1),
                "message": f"Welcome back, {username}!"
            })

        return jsonify({"match": None, "confidence": 0, "message": "Face not recognized"})

    @app.route("/api/user/face-status", methods=["GET"])
    @_login_required
    def face_status():
        """Check the current user's face enrollment status."""
        face_mgr = app.config.get("FACE_PROFILE_MANAGER")
        if not face_mgr:
            return jsonify({"enrolled": False, "angles": []})
        username = _get_current_username()
        profile = face_mgr.get_face_profile(username)
        return jsonify({
            "enrolled": face_mgr.is_face_enrolled(username),
            "angles": list(profile.keys()),
            "remaining": list({"front", "up", "down", "left", "right"} - set(profile.keys())),
        })

    # ── Knowledge Base API ─────────────────────────────────────────────────

    @app.route("/api/kb/documents", methods=["GET"])
    @_login_required
    def kb_list_documents():
        """List knowledge base documents (user sees own, admin sees all)."""
        kb = app.config.get("KB_MANAGER")
        if not kb:
            return jsonify([])
        username = _get_current_username()
        perms = _get_user_permissions()
        include_all = "admin" in perms
        return jsonify(kb.get_documents(username, include_all=include_all))

    @app.route("/api/kb/documents", methods=["POST"])
    @_login_required
    def kb_add_document():
        """Add a document to the knowledge base."""
        kb = app.config.get("KB_MANAGER")
        if not kb:
            return jsonify({"error": "KB not available"}), 500
        data = request.get_json(silent=True)
        if not data or "title" not in data or "content" not in data:
            return jsonify({"error": "Missing 'title' and 'content'"}), 400
        username = _get_current_username()
        result = kb.add_document(username, data["title"], data["content"], data.get("source", "manual"))
        return jsonify(result)

    @app.route("/api/kb/documents/<int:doc_id>", methods=["GET"])
    @_login_required
    def kb_get_document(doc_id):
        """Get a document with its content."""
        kb = app.config.get("KB_MANAGER")
        if not kb:
            return jsonify({"error": "KB not available"}), 500
        doc = kb.get_document(doc_id)
        if not doc:
            return jsonify({"error": "Document not found"}), 404
        return jsonify(doc)

    @app.route("/api/kb/documents/<int:doc_id>", methods=["PUT"])
    @_login_required
    def kb_update_document(doc_id):
        """Update a document (re-chunks if content changes)."""
        kb = app.config.get("KB_MANAGER")
        if not kb:
            return jsonify({"error": "KB not available"}), 500
        data = request.get_json(silent=True) or {}
        success = kb.update_document(doc_id, title=data.get("title"), content=data.get("content"))
        if success:
            return jsonify({"success": True, "message": "Document updated"})
        return jsonify({"error": "Document not found"}), 404

    @app.route("/api/kb/documents/<int:doc_id>", methods=["DELETE"])
    @_login_required
    def kb_delete_document(doc_id):
        """Delete a document and all its chunks."""
        kb = app.config.get("KB_MANAGER")
        if not kb:
            return jsonify({"error": "KB not available"}), 500
        if kb.delete_document(doc_id):
            return jsonify({"success": True, "message": "Document deleted"})
        return jsonify({"error": "Document not found"}), 404

    @app.route("/api/kb/documents/<int:doc_id>/chunks", methods=["GET"])
    @_login_required
    def kb_get_chunks(doc_id):
        """Get all chunks for a document."""
        kb = app.config.get("KB_MANAGER")
        if not kb:
            return jsonify([])
        return jsonify(kb.get_chunks(doc_id))

    @app.route("/api/kb/chunks/<int:chunk_id>", methods=["PUT"])
    @_login_required
    def kb_update_chunk(chunk_id):
        """Update a chunk's content."""
        kb = app.config.get("KB_MANAGER")
        if not kb:
            return jsonify({"error": "KB not available"}), 500
        data = request.get_json(silent=True)
        if not data or "content" not in data:
            return jsonify({"error": "Missing 'content'"}), 400
        if kb.update_chunk(chunk_id, data["content"]):
            return jsonify({"success": True})
        return jsonify({"error": "Chunk not found"}), 404

    @app.route("/api/kb/chunks/<int:chunk_id>", methods=["DELETE"])
    @_login_required
    def kb_delete_chunk(chunk_id):
        """Delete a specific chunk."""
        kb = app.config.get("KB_MANAGER")
        if not kb:
            return jsonify({"error": "KB not available"}), 500
        if kb.delete_chunk(chunk_id):
            return jsonify({"success": True})
        return jsonify({"error": "Chunk not found"}), 404

    @app.route("/api/kb/search", methods=["GET"])
    @_login_required
    def kb_search():
        """Search the knowledge base."""
        kb = app.config.get("KB_MANAGER")
        if not kb:
            return jsonify([])
        query = request.args.get("q", "")
        username = _get_current_username()
        return jsonify(kb.search(username, query))

    @app.route("/knowledge", methods=["GET"])
    @_login_required
    def knowledge_page():
        """Knowledge Base management UI page."""
        return render_template("knowledge.html")

    @app.route("/help", methods=["GET"])
    @_login_required
    def help_page():
        """Help & FAQ page."""
        return render_template("help.html")

    # --- Daily Briefing API ---

    @app.route("/api/briefing", methods=["GET"])
    @_login_required
    def get_briefing():
        """Generate and return the daily briefing for the current user."""
        briefing = app.config.get("DAILY_BRIEFING")
        if not briefing:
            return jsonify({"error": "Briefing not available"}), 500
        username = _get_current_username()
        # Get user honorific
        user_prefs = app.config.get("USER_PREFS_MANAGER")
        honorific = "Sir"
        if user_prefs and username:
            prefs = user_prefs.get_preferences(username)
            honorific = prefs.get("honorific", "Sir") or "Sir"
        text = briefing.generate(username, honorific)
        return jsonify({"briefing": text})

    @app.route("/api/briefing/settings", methods=["GET"])
    @_login_required
    def get_briefing_settings():
        """Get briefing settings for the current user."""
        briefing = app.config.get("DAILY_BRIEFING")
        if not briefing:
            return jsonify({"error": "Briefing not available"}), 500
        username = _get_current_username()
        settings = briefing.get_settings(username)
        return jsonify(settings)

    @app.route("/api/briefing/settings", methods=["PUT"])
    @_login_required
    def update_briefing_settings():
        """Update briefing settings for the current user."""
        briefing = app.config.get("DAILY_BRIEFING")
        if not briefing:
            return jsonify({"error": "Briefing not available"}), 500
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "No data provided"}), 400
        username = _get_current_username()
        if briefing.save_settings(username, data):
            return jsonify({"success": True, "message": "Briefing settings saved."})
        return jsonify({"error": "Failed to save settings"}), 500

    # --- Workflow Automation API ---

    @app.route("/api/workflows", methods=["GET"])
    @_login_required
    def list_workflows():
        """List all workflows for the current user (admins see all)."""
        engine = app.config.get("WORKFLOW_ENGINE")
        if not engine:
            return jsonify([])
        perms = _get_user_permissions()
        username = _get_current_username()
        if "admin" in perms:
            workflows = engine.list_workflows(include_disabled=True)
        else:
            workflows = engine.list_workflows(username=username, include_disabled=True)
        return jsonify(workflows)

    @app.route("/api/workflows", methods=["POST"])
    @_login_required
    def create_workflow():
        """Create a new workflow."""
        engine = app.config.get("WORKFLOW_ENGINE")
        if not engine:
            return jsonify({"error": "Workflow engine not available"}), 500
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "No data provided"}), 400

        required = ["name", "trigger_type", "trigger_config", "action_type", "action_config"]
        for field in required:
            if field not in data:
                return jsonify({"error": f"Missing required field: '{field}'"}), 400

        username = _get_current_username()
        result = engine.create_workflow(
            name=data["name"],
            trigger_type=data["trigger_type"],
            trigger_config=data["trigger_config"],
            action_type=data["action_type"],
            action_config=data["action_config"],
            description=data.get("description", ""),
            conditions=data.get("conditions"),
            username=username,
        )
        if "error" in result:
            return jsonify(result), 400
        return jsonify(result)

    @app.route("/api/workflows/<int:workflow_id>", methods=["GET"])
    @_login_required
    def get_workflow(workflow_id):
        """Get a single workflow by ID."""
        engine = app.config.get("WORKFLOW_ENGINE")
        if not engine:
            return jsonify({"error": "Workflow engine not available"}), 500
        wf = engine.get_workflow(workflow_id)
        if not wf:
            return jsonify({"error": "Workflow not found"}), 404
        return jsonify(wf)

    @app.route("/api/workflows/<int:workflow_id>", methods=["PUT"])
    @_login_required
    def update_workflow(workflow_id):
        """Update a workflow."""
        engine = app.config.get("WORKFLOW_ENGINE")
        if not engine:
            return jsonify({"error": "Workflow engine not available"}), 500
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "No data provided"}), 400
        result = engine.update_workflow(workflow_id, data)
        if "error" in result:
            return jsonify(result), 400
        return jsonify(result)

    @app.route("/api/workflows/<int:workflow_id>", methods=["DELETE"])
    @_login_required
    def delete_workflow(workflow_id):
        """Delete a workflow."""
        engine = app.config.get("WORKFLOW_ENGINE")
        if not engine:
            return jsonify({"error": "Workflow engine not available"}), 500
        result = engine.delete_workflow(workflow_id)
        if "error" in result:
            return jsonify(result), 404
        return jsonify(result)

    @app.route("/api/workflows/<int:workflow_id>/logs", methods=["GET"])
    @_login_required
    def get_workflow_logs(workflow_id):
        """Get execution logs for a workflow."""
        engine = app.config.get("WORKFLOW_ENGINE")
        if not engine:
            return jsonify([])
        limit = request.args.get("limit", 20, type=int)
        return jsonify(engine.get_workflow_logs(workflow_id, limit=limit))

    @app.route("/api/workflows/<int:workflow_id>/test", methods=["POST"])
    @_login_required
    def test_workflow(workflow_id):
        """Manually trigger a workflow for testing."""
        engine = app.config.get("WORKFLOW_ENGINE")
        if not engine:
            return jsonify({"error": "Workflow engine not available"}), 500
        wf = engine.get_workflow(workflow_id)
        if not wf:
            return jsonify({"error": "Workflow not found"}), 404
        engine._execute_action(wf)
        return jsonify({"message": f"Workflow '{wf['name']}' triggered manually."})

    # --- Code Sandbox API ---

    @app.route("/api/sandbox/run", methods=["POST"])
    @_permission_required("commands")
    def run_sandbox_code():
        """Execute Python code in the sandbox."""
        from app.code_sandbox import CodeSandbox

        data = request.get_json(silent=True)
        if not data or "code" not in data:
            return jsonify({"error": "Missing 'code' field"}), 400

        timeout = data.get("timeout", 10)
        if timeout > 30:
            timeout = 30  # Cap at 30 seconds

        sandbox = CodeSandbox(timeout=timeout, allow_imports=True)
        result = sandbox.execute(data["code"])
        return jsonify(result)

    # --- Contextual Suggestions API ---

    @app.route("/api/suggestions", methods=["GET"])
    @_login_required
    def get_suggestions():
        """Get current suggestions for the user."""
        engine = app.config.get("SUGGESTIONS_ENGINE")
        if not engine:
            return jsonify([])
        username = _get_current_username()
        suggestions = engine.analyze_patterns(username)
        return jsonify(suggestions[:5])

    @app.route("/api/suggestions/history", methods=["GET"])
    @_login_required
    def get_suggestion_history():
        """Get past suggestions for the user."""
        engine = app.config.get("SUGGESTIONS_ENGINE")
        if not engine:
            return jsonify([])
        username = _get_current_username()
        limit = request.args.get("limit", 20, type=int)
        return jsonify(engine.get_suggestion_history(username, limit=limit))

    @app.route("/api/suggestions/<int:suggestion_id>/dismiss", methods=["POST"])
    @_login_required
    def dismiss_suggestion(suggestion_id):
        """Dismiss a suggestion."""
        engine = app.config.get("SUGGESTIONS_ENGINE")
        if not engine:
            return jsonify({"error": "Not available"}), 500
        engine.dismiss_suggestion(suggestion_id)
        return jsonify({"success": True})

    @app.route("/api/suggestions/<int:suggestion_id>/accept", methods=["POST"])
    @_login_required
    def accept_suggestion(suggestion_id):
        """Accept a suggestion."""
        engine = app.config.get("SUGGESTIONS_ENGINE")
        if not engine:
            return jsonify({"error": "Not available"}), 500
        engine.accept_suggestion(suggestion_id)
        return jsonify({"success": True})

    @app.route("/api/suggestions/activity", methods=["GET"])
    @_login_required
    def get_activity_stats():
        """Get activity statistics for the current user."""
        engine = app.config.get("SUGGESTIONS_ENGINE")
        if not engine:
            return jsonify({})
        username = _get_current_username()
        return jsonify(engine.get_activity_stats(username))

    @app.route("/api/suggestions/ingest-history", methods=["POST"])
    @_login_required
    def ingest_os_history():
        """Ingest OS command history for pattern analysis."""
        engine = app.config.get("SUGGESTIONS_ENGINE")
        if not engine:
            return jsonify({"error": "Not available"}), 500
        username = _get_current_username()
        result = engine.ingest_os_history(username)
        return jsonify(result)

    # --- Cron Job Manager API ---

    @app.route("/api/cron", methods=["GET"])
    @_permission_required("commands")
    def list_cron_jobs():
        """List all cron jobs."""
        cron = app.config.get("CRON_MANAGER")
        if not cron:
            return jsonify([])
        perms = _get_user_permissions()
        username = _get_current_username() if "admin" not in perms else None
        return jsonify(cron.list_jobs(username))

    @app.route("/api/cron", methods=["POST"])
    @_permission_required("commands")
    def create_cron_job():
        """Create a new cron job."""
        cron = app.config.get("CRON_MANAGER")
        if not cron:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        username = _get_current_username()
        result = cron.create_job(
            name=data.get("name", ""), command=data.get("command", ""),
            schedule=data.get("schedule", ""), description=data.get("description", ""),
            username=username)
        return jsonify(result), 200 if "id" in result else 400

    @app.route("/api/cron/<int:job_id>", methods=["GET"])
    @_permission_required("commands")
    def get_cron_job(job_id):
        cron = app.config.get("CRON_MANAGER")
        if not cron:
            return jsonify({"error": "Not available"}), 500
        job = cron.get_job(job_id)
        return jsonify(job) if job else (jsonify({"error": "Not found"}), 404)

    @app.route("/api/cron/<int:job_id>", methods=["PUT"])
    @_permission_required("commands")
    def update_cron_job(job_id):
        cron = app.config.get("CRON_MANAGER")
        if not cron:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        return jsonify(cron.update_job(job_id, data))

    @app.route("/api/cron/<int:job_id>", methods=["DELETE"])
    @_permission_required("commands")
    def delete_cron_job(job_id):
        cron = app.config.get("CRON_MANAGER")
        if not cron:
            return jsonify({"error": "Not available"}), 500
        return jsonify(cron.delete_job(job_id))

    @app.route("/api/cron/<int:job_id>/run", methods=["POST"])
    @_permission_required("commands")
    def run_cron_job(job_id):
        """Manually trigger a cron job."""
        cron = app.config.get("CRON_MANAGER")
        if not cron:
            return jsonify({"error": "Not available"}), 500
        return jsonify(cron.run_job(job_id))

    @app.route("/api/cron/<int:job_id>/history", methods=["GET"])
    @_permission_required("commands")
    def get_cron_history(job_id):
        cron = app.config.get("CRON_MANAGER")
        if not cron:
            return jsonify([])
        limit = request.args.get("limit", 20, type=int)
        return jsonify(cron.get_history(job_id, limit))

    # --- Log Analyzer API ---

    @app.route("/api/logs/watches", methods=["GET"])
    @_permission_required("monitoring")
    def list_log_watches():
        analyzer = app.config.get("LOG_ANALYZER")
        return jsonify(analyzer.list_watches() if analyzer else [])

    @app.route("/api/logs/watches", methods=["POST"])
    @_permission_required("monitoring")
    def add_log_watch():
        analyzer = app.config.get("LOG_ANALYZER")
        if not analyzer:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        username = _get_current_username()
        result = analyzer.add_watch(
            file_path=data.get("file_path", ""),
            label=data.get("label"), error_threshold=data.get("error_threshold", 5),
            window_minutes=data.get("window_minutes", 5), username=username)
        return jsonify(result), 200 if "message" in result else 400

    @app.route("/api/logs/tail", methods=["GET"])
    @_permission_required("monitoring")
    def tail_log():
        analyzer = app.config.get("LOG_ANALYZER")
        if not analyzer:
            return jsonify({"error": "Not available"}), 500
        file_path = request.args.get("file", "")
        lines = request.args.get("lines", 50, type=int)
        return jsonify(analyzer.tail(file_path, lines))

    @app.route("/api/logs/search", methods=["GET"])
    @_permission_required("monitoring")
    def search_log():
        analyzer = app.config.get("LOG_ANALYZER")
        if not analyzer:
            return jsonify({"error": "Not available"}), 500
        file_path = request.args.get("file", "")
        pattern = request.args.get("pattern", "")
        return jsonify(analyzer.search(file_path, pattern))

    @app.route("/api/logs/errors", methods=["GET"])
    @_permission_required("monitoring")
    def get_log_errors():
        analyzer = app.config.get("LOG_ANALYZER")
        if not analyzer:
            return jsonify({"error": "Not available"}), 500
        file_path = request.args.get("file", "")
        return jsonify(analyzer.get_error_summary(file_path))

    @app.route("/api/logs/alerts", methods=["GET"])
    @_permission_required("monitoring")
    def get_log_alerts():
        analyzer = app.config.get("LOG_ANALYZER")
        if not analyzer:
            return jsonify([])
        unacked = request.args.get("unacknowledged", "false") == "true"
        return jsonify(analyzer.get_alerts(unacknowledged_only=unacked))

    @app.route("/api/logs/alerts/<int:alert_id>/acknowledge", methods=["POST"])
    @_permission_required("monitoring")
    def acknowledge_log_alert(alert_id):
        analyzer = app.config.get("LOG_ANALYZER")
        if not analyzer:
            return jsonify({"error": "Not available"}), 500
        analyzer.acknowledge_alert(alert_id)
        return jsonify({"success": True})

    # --- Service Manager API ---

    @app.route("/api/services", methods=["GET"])
    @_permission_required("commands")
    def list_services():
        sm = app.config.get("SERVICE_MANAGER")
        if not sm:
            return jsonify({"error": "Not available"}), 500
        host = request.args.get("host")
        return jsonify(sm.service_list(host=host))

    @app.route("/api/services/<service_name>/<action>", methods=["POST"])
    @_permission_required("admin")
    def manage_service(service_name, action):
        sm = app.config.get("SERVICE_MANAGER")
        if not sm:
            return jsonify({"error": "Not available"}), 500
        host = (request.get_json(silent=True) or {}).get("host")
        return jsonify(sm.service_action(service_name, action, host=host))

    @app.route("/api/services/<service_name>/status", methods=["GET"])
    @_permission_required("commands")
    def get_service_status(service_name):
        sm = app.config.get("SERVICE_MANAGER")
        if not sm:
            return jsonify({"error": "Not available"}), 500
        host = request.args.get("host")
        return jsonify(sm.service_status(service_name, host=host))

    @app.route("/api/docker/containers", methods=["GET"])
    @_permission_required("commands")
    def list_docker_containers():
        sm = app.config.get("SERVICE_MANAGER")
        if not sm:
            return jsonify({"error": "Not available"}), 500
        all_c = request.args.get("all", "false") == "true"
        host = request.args.get("host")
        return jsonify(sm.docker_list(all_containers=all_c, host=host))

    @app.route("/api/docker/containers/<container>/<action>", methods=["POST"])
    @_permission_required("admin")
    def manage_docker_container(container, action):
        sm = app.config.get("SERVICE_MANAGER")
        if not sm:
            return jsonify({"error": "Not available"}), 500
        host = (request.get_json(silent=True) or {}).get("host")
        return jsonify(sm.docker_action(container, action, host=host))

    @app.route("/api/docker/containers/<container>/logs", methods=["GET"])
    @_permission_required("commands")
    def get_docker_logs(container):
        sm = app.config.get("SERVICE_MANAGER")
        if not sm:
            return jsonify({"error": "Not available"}), 500
        lines = request.args.get("lines", 50, type=int)
        host = request.args.get("host")
        return jsonify(sm.docker_logs(container, lines=lines, host=host))

    @app.route("/api/docker/stats", methods=["GET"])
    @_permission_required("monitoring")
    def get_docker_stats():
        sm = app.config.get("SERVICE_MANAGER")
        if not sm:
            return jsonify({"error": "Not available"}), 500
        host = request.args.get("host")
        return jsonify(sm.docker_stats(host=host))

    @app.route("/api/git/status", methods=["GET"])
    @_permission_required("commands")
    def get_git_status():
        sm = app.config.get("SERVICE_MANAGER")
        if not sm:
            return jsonify({"error": "Not available"}), 500
        repo = request.args.get("repo", ".")
        host = request.args.get("host")
        return jsonify(sm.git_status(repo_path=repo, host=host))

    @app.route("/api/git/pull", methods=["POST"])
    @_permission_required("admin")
    def do_git_pull():
        sm = app.config.get("SERVICE_MANAGER")
        if not sm:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        return jsonify(sm.git_pull(repo_path=data.get("repo", "."), host=data.get("host")))

    @app.route("/api/git/log", methods=["GET"])
    @_permission_required("commands")
    def get_git_log():
        sm = app.config.get("SERVICE_MANAGER")
        if not sm:
            return jsonify({"error": "Not available"}), 500
        repo = request.args.get("repo", ".")
        count = request.args.get("count", 10, type=int)
        host = request.args.get("host")
        return jsonify(sm.git_log(repo_path=repo, count=count, host=host))

    # --- Backup Orchestrator API ---

    @app.route("/api/backups", methods=["GET"])
    @_permission_required("admin")
    def list_backups():
        bo = app.config.get("BACKUP_ORCHESTRATOR")
        if not bo:
            return jsonify([])
        return jsonify(bo.get_history())

    @app.route("/api/backups/run", methods=["POST"])
    @_permission_required("admin")
    def trigger_backup():
        bo = app.config.get("BACKUP_ORCHESTRATOR")
        if not bo:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        backup_type = data.get("type", "full")
        return jsonify(bo.run_backup(backup_type))

    @app.route("/api/backups/config", methods=["GET"])
    @_permission_required("admin")
    def get_backup_config():
        bo = app.config.get("BACKUP_ORCHESTRATOR")
        if not bo:
            return jsonify({})
        return jsonify(bo.get_config())

    @app.route("/api/backups/config", methods=["PUT"])
    @_permission_required("admin")
    def update_backup_config():
        bo = app.config.get("BACKUP_ORCHESTRATOR")
        if not bo:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        return jsonify(bo.save_config(data))

    @app.route("/api/backups/<int:backup_id>/verify", methods=["POST"])
    @_permission_required("admin")
    def verify_backup_integrity(backup_id):
        bo = app.config.get("BACKUP_ORCHESTRATOR")
        if not bo:
            return jsonify({"error": "Not available"}), 500
        return jsonify(bo.verify_backup(backup_id))

    @app.route("/api/backups/heartbeat", methods=["GET"])
    def check_heartbeat():
        """Public endpoint for external monitoring to check if Jarvis is alive."""
        bo = app.config.get("BACKUP_ORCHESTRATOR")
        if not bo:
            return jsonify({"status": "unknown"})
        return jsonify(bo.get_heartbeat())

    @app.route("/api/backups/dr-plan", methods=["GET"])
    @_permission_required("admin")
    def get_disaster_recovery_plan():
        """Generate disaster recovery plan."""
        bo = app.config.get("BACKUP_ORCHESTRATOR")
        if not bo:
            return jsonify({"error": "Not available"}), 500
        return jsonify(bo.generate_dr_plan())

    # --- WiFi Network Manager API ---

    @app.route("/wifi", methods=["GET"])
    @_permission_required("bluetooth")
    def wifi_page():
        """WiFi network management UI page."""
        return render_template("wifi.html")

    @app.route("/assistants", methods=["GET"])
    @_login_required
    def assistants_page():
        """Custom Assistants management UI page."""
        return render_template("assistants.html")

    @app.route("/api/wifi/devices", methods=["GET"])
    @_permission_required("bluetooth")
    def list_wifi_devices():
        """List all known WiFi devices."""
        wm = app.config.get("WIFI_MANAGER")
        if not wm:
            return jsonify([])
        return jsonify(wm.get_devices())

    @app.route("/api/wifi/scan", methods=["POST"])
    @_permission_required("bluetooth")
    def scan_wifi_network():
        """Scan the network for connected devices."""
        wm = app.config.get("WIFI_MANAGER")
        if not wm:
            return jsonify({"error": "WiFi manager not available"}), 500
        devices = wm.scan_network()
        return jsonify({"devices": devices, "count": len(devices)})

    @app.route("/api/wifi/devices/<mac>", methods=["PUT"])
    @_permission_required("bluetooth")
    def update_wifi_device(mac):
        """Update device metadata (name, SSH credentials, notes)."""
        wm = app.config.get("WIFI_MANAGER")
        if not wm:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        return jsonify(wm.update_device(mac, data))

    @app.route("/api/wifi/ssh/connect", methods=["POST"])
    @_permission_required("commands")
    def wifi_ssh_connect():
        """Attempt SSH connection to a WiFi device."""
        wm = app.config.get("WIFI_MANAGER")
        if not wm:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        ip = data.get("ip", "")
        if not ip:
            return jsonify({"error": "Missing 'ip' field"}), 400
        result = wm.ssh_connect(
            ip=ip,
            username=data.get("username"),
            password=data.get("password"),
            port=data.get("port", 22),
            auto_try=data.get("auto_try", True),
        )
        status_code = 200 if result.get("success") else 401 if result.get("needs_credentials") else 500
        return jsonify(result), status_code

    @app.route("/api/wifi/ssh/disconnect", methods=["POST"])
    @_permission_required("commands")
    def wifi_ssh_disconnect():
        """Disconnect an active SSH session."""
        wm = app.config.get("WIFI_MANAGER")
        if not wm:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        ip = data.get("ip", "")
        if not ip:
            return jsonify({"error": "Missing 'ip' field"}), 400
        return jsonify(wm.ssh_disconnect(ip))

    @app.route("/api/wifi/ssh/execute", methods=["POST"])
    @_permission_required("commands")
    def wifi_ssh_execute():
        """Execute a command on a connected device via SSH."""
        wm = app.config.get("WIFI_MANAGER")
        if not wm:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        ip = data.get("ip", "")
        command = data.get("command", "")
        if not ip or not command:
            return jsonify({"error": "Missing 'ip' and 'command' fields"}), 400
        return jsonify(wm.ssh_execute(ip, command, timeout=data.get("timeout", 30)))

    @app.route("/api/wifi/sessions", methods=["GET"])
    @_permission_required("commands")
    def list_wifi_ssh_sessions():
        """List active SSH sessions."""
        wm = app.config.get("WIFI_MANAGER")
        if not wm:
            return jsonify([])
        return jsonify(wm.get_active_sessions())

    # --- Custom Assistants API ---

    @app.route("/api/assistants", methods=["GET"])
    @_login_required
    def list_assistants():
        """List all assistants available to the current user."""
        cam = app.config.get("CUSTOM_ASSISTANTS")
        if not cam:
            return jsonify([])
        username = _get_current_username()
        return jsonify(cam.list_assistants(username))

    @app.route("/api/assistants", methods=["POST"])
    @_login_required
    def create_assistant():
        """Create a new custom assistant."""
        cam = app.config.get("CUSTOM_ASSISTANTS")
        if not cam:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        username = _get_current_username()
        result = cam.create_assistant(
            name=data.get("name", ""),
            system_prompt=data.get("system_prompt", ""),
            username=username,
            description=data.get("description", ""),
            inference_params=data.get("inference_params"),
            icon=data.get("icon", "🤖"),
            shared=data.get("shared", False),
        )
        return jsonify(result), 200 if "id" in result else 400

    @app.route("/api/assistants/<int:assistant_id>", methods=["GET"])
    @_login_required
    def get_assistant_detail(assistant_id):
        """Get a single assistant's full configuration."""
        cam = app.config.get("CUSTOM_ASSISTANTS")
        if not cam:
            return jsonify({"error": "Not available"}), 500
        assistant = cam.get_assistant(assistant_id)
        if not assistant:
            return jsonify({"error": "Not found"}), 404
        return jsonify(assistant)

    @app.route("/api/assistants/<int:assistant_id>", methods=["PUT"])
    @_login_required
    def update_assistant_config(assistant_id):
        """Update a custom assistant (owner or admin)."""
        cam = app.config.get("CUSTOM_ASSISTANTS")
        if not cam:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        username = _get_current_username()
        perms = _get_user_permissions()
        is_admin = "admin" in perms
        result = cam.update_assistant(assistant_id, data, username, is_admin=is_admin)
        return jsonify(result), 200 if "message" in result else 400

    @app.route("/api/assistants/<int:assistant_id>", methods=["DELETE"])
    @_login_required
    def delete_assistant_config(assistant_id):
        """Delete a custom assistant (owner or admin)."""
        cam = app.config.get("CUSTOM_ASSISTANTS")
        if not cam:
            return jsonify({"error": "Not available"}), 500
        username = _get_current_username()
        perms = _get_user_permissions()
        is_admin = "admin" in perms
        result = cam.delete_assistant(assistant_id, username, is_admin=is_admin)
        return jsonify(result), 200 if "message" in result else 400

    @app.route("/api/assistants/default-prompt", methods=["GET"])
    @_login_required
    def get_default_prompt():
        """Get the default Jarvis system prompt as a template for new assistants."""
        cam = app.config.get("CUSTOM_ASSISTANTS")
        if not cam:
            return jsonify({"error": "Not available"}), 500
        return jsonify({"prompt": cam.get_default_system_prompt()})

    # --- Flow Engine API (Visual Workflows) ---

    @app.route("/workflows", methods=["GET"])
    @_login_required
    def workflows_page():
        """Visual workflow builder UI page."""
        return render_template("workflows.html")

    @app.route("/api/flows/block-types", methods=["GET"])
    @_login_required
    def get_flow_block_types():
        """Get available block types for the flow editor."""
        fe = app.config.get("FLOW_ENGINE")
        if not fe:
            return jsonify({})
        return jsonify(fe.get_block_types())

    @app.route("/api/flows", methods=["GET"])
    @_login_required
    def list_flows():
        """List flows for the current user (admins can see all with ?all=true)."""
        fe = app.config.get("FLOW_ENGINE")
        if not fe:
            return jsonify([])
        perms = _get_user_permissions()
        username = _get_current_username()
        include_all = request.args.get("all", "false") == "true" and "admin" in perms
        return jsonify(fe.list_flows(username=username, include_all=include_all))

    @app.route("/api/flows", methods=["POST"])
    @_login_required
    def create_flow():
        """Create a new flow."""
        fe = app.config.get("FLOW_ENGINE")
        if not fe:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        username = _get_current_username()
        result = fe.create_flow(
            name=data.get("name", ""),
            username=username,
            description=data.get("description", ""),
            blocks=data.get("blocks"),
            connections=data.get("connections"),
            schedule=data.get("schedule"),
        )
        return jsonify(result), 200 if "id" in result else 400

    @app.route("/api/flows/<int:flow_id>", methods=["GET"])
    @_login_required
    def get_flow_detail(flow_id):
        """Get a flow's full configuration."""
        fe = app.config.get("FLOW_ENGINE")
        if not fe:
            return jsonify({"error": "Not available"}), 500
        flow = fe.get_flow(flow_id)
        if not flow:
            return jsonify({"error": "Not found"}), 404
        return jsonify(flow)

    @app.route("/api/flows/<int:flow_id>", methods=["PUT"])
    @_login_required
    def update_flow(flow_id):
        """Update a flow."""
        fe = app.config.get("FLOW_ENGINE")
        if not fe:
            return jsonify({"error": "Not available"}), 500
        data = request.get_json(silent=True) or {}
        return jsonify(fe.update_flow(flow_id, data))

    @app.route("/api/flows/<int:flow_id>", methods=["DELETE"])
    @_login_required
    def delete_flow_route(flow_id):
        """Delete a flow and its history."""
        fe = app.config.get("FLOW_ENGINE")
        if not fe:
            return jsonify({"error": "Not available"}), 500
        return jsonify(fe.delete_flow(flow_id))

    @app.route("/api/flows/<int:flow_id>/run", methods=["POST"])
    @_login_required
    def run_flow_route(flow_id):
        """Execute a flow manually."""
        fe = app.config.get("FLOW_ENGINE")
        if not fe:
            return jsonify({"error": "Not available"}), 500
        return jsonify(fe.execute_flow(flow_id, triggered_by="manual"))

    @app.route("/api/flows/<int:flow_id>/runs", methods=["GET"])
    @_login_required
    def get_flow_runs(flow_id):
        """Get execution history for a flow."""
        fe = app.config.get("FLOW_ENGINE")
        if not fe:
            return jsonify([])
        limit = request.args.get("limit", 20, type=int)
        return jsonify(fe.get_runs(flow_id, limit=limit))
