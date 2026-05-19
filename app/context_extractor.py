"""Context Extractor for Jarvis Assistant.

Extracts text content from various sources:
- Uploaded files (PDF, TXT, DOCX, MD, CSV, JSON, code files)
- Web URLs (fetches and extracts readable text)

Designed to be lightweight for the Raspberry Pi.
"""

import logging
import re
from pathlib import Path
from typing import Optional

import requests

logger = logging.getLogger(__name__)

REQUEST_TIMEOUT = 15
MAX_CONTENT_LENGTH = 8000  # Max chars to inject as context (keep within token limits)

# File extensions we can handle
SUPPORTED_EXTENSIONS = {
    ".txt", ".md", ".csv", ".json", ".xml", ".yaml", ".yml",
    ".py", ".js", ".ts", ".html", ".css", ".sh", ".bash",
    ".c", ".cpp", ".h", ".java", ".rs", ".go", ".rb",
    ".log", ".conf", ".ini", ".toml", ".env",
    ".pdf", ".docx",
    ".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".gif", ".webp",
    ".xlsx", ".xls", ".ods",
}


def extract_from_file(file_storage) -> dict:
    """Extract text content from an uploaded file.

    Args:
        file_storage: Flask FileStorage object from request.files.

    Returns:
        Dict with 'content' (str) and 'filename' (str), or 'error' (str).
    """
    filename = file_storage.filename or "unknown"
    ext = Path(filename).suffix.lower()

    if ext not in SUPPORTED_EXTENSIONS:
        # Try to read as plain text for unknown extensions
        try:
            raw = file_storage.read()
            content = raw.decode("utf-8", errors="replace")
            if len(content) > MAX_CONTENT_LENGTH:
                content = content[:MAX_CONTENT_LENGTH] + f"\n\n[... truncated, {len(content)} total chars]"
            return {"content": content, "filename": filename}
        except Exception:
            return {"error": f"Unsupported file type: {ext}. Could not read as text."}

    try:
        if ext == ".pdf":
            content = _extract_pdf(file_storage)
        elif ext == ".docx":
            content = _extract_docx(file_storage)
        elif ext in (".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".gif", ".webp"):
            content = _extract_image_ocr(file_storage)
        elif ext in (".xlsx", ".xls", ".ods"):
            content = _extract_excel(file_storage)
        else:
            # Plain text files (code, markdown, config, etc.)
            raw = file_storage.read()
            content = raw.decode("utf-8", errors="replace")

        # Truncate if too long
        if len(content) > MAX_CONTENT_LENGTH:
            content = content[:MAX_CONTENT_LENGTH] + f"\n\n[... truncated, {len(content)} total chars]"

        return {"content": content, "filename": filename}

    except Exception as e:
        logger.error("File extraction error for '%s': %s", filename, e)
        return {"error": f"Failed to extract content from '{filename}': {e}"}


def extract_from_url(url: str) -> dict:
    """Fetch and extract readable text from a web URL.

    Args:
        url: The URL to fetch.

    Returns:
        Dict with 'content' (str) and 'url' (str), or 'error' (str).
    """
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    try:
        resp = requests.get(
            url,
            timeout=REQUEST_TIMEOUT,
            headers={"User-Agent": "JarvisAssistant/1.0 (Raspberry Pi)"},
        )

        if resp.status_code >= 400:
            return {"error": f"Failed to fetch URL (HTTP {resp.status_code}): {url}"}

        content_type = resp.headers.get("content-type", "")

        if "text/html" in content_type:
            content = _extract_html(resp.text)
        elif "application/json" in content_type:
            content = resp.text[:MAX_CONTENT_LENGTH]
        elif "text/" in content_type:
            content = resp.text[:MAX_CONTENT_LENGTH]
        else:
            return {"error": f"Unsupported content type: {content_type}"}

        if len(content) > MAX_CONTENT_LENGTH:
            content = content[:MAX_CONTENT_LENGTH] + f"\n\n[... truncated]"

        return {"content": content, "url": url}

    except requests.exceptions.Timeout:
        return {"error": f"Timeout fetching URL: {url}"}
    except requests.exceptions.ConnectionError:
        return {"error": f"Could not connect to: {url}"}
    except Exception as e:
        logger.error("URL extraction error for '%s': %s", url, e)
        return {"error": f"Failed to fetch URL: {e}"}


def _extract_html(html: str) -> str:
    """Extract readable text from HTML, stripping tags and scripts."""
    # Remove script and style blocks
    text = re.sub(r"<script[^>]*>.*?</script>", "", html, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL | re.IGNORECASE)
    # Remove HTML comments
    text = re.sub(r"<!--.*?-->", "", text, flags=re.DOTALL)
    # Remove nav, header, footer
    text = re.sub(r"<(nav|header|footer)[^>]*>.*?</\1>", "", text, flags=re.DOTALL | re.IGNORECASE)
    # Strip remaining tags
    text = re.sub(r"<[^>]+>", " ", text)
    # Decode common entities
    text = text.replace("&nbsp;", " ").replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">").replace("&quot;", '"')
    # Collapse whitespace
    text = re.sub(r"\s+", " ", text).strip()
    # Try to preserve some structure with newlines at block boundaries
    text = re.sub(r"\s{3,}", "\n\n", text)
    return text


def _extract_pdf(file_storage) -> str:
    """Extract text from a PDF file. Uses PyPDF2 if available, else basic fallback."""
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(file_storage)
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n\n".join(pages) if pages else "(PDF contained no extractable text)"
    except ImportError:
        return "(PDF support requires PyPDF2: pip install PyPDF2)"
    except Exception as e:
        return f"(PDF extraction failed: {e})"


def _extract_docx(file_storage) -> str:
    """Extract text from a DOCX file. Uses python-docx if available."""
    try:
        import docx
        doc = docx.Document(file_storage)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs) if paragraphs else "(DOCX contained no text)"
    except ImportError:
        return "(DOCX support requires python-docx: pip install python-docx)"
    except Exception as e:
        return f"(DOCX extraction failed: {e})"


def _extract_image_ocr(file_storage) -> str:
    """Extract text from an image using OCR (pytesseract or easyocr)."""
    # Try pytesseract first (requires Tesseract installed on the system)
    try:
        from PIL import Image
        import pytesseract

        # Auto-detect Tesseract on Windows if not in PATH
        import platform
        if platform.system() == "Windows":
            import os
            common_paths = [
                r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                os.path.expanduser(r"~\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"),
            ]
            for p in common_paths:
                if os.path.isfile(p):
                    pytesseract.pytesseract.tesseract_cmd = p
                    break

        file_storage.seek(0)
        image = Image.open(file_storage)
        text = pytesseract.image_to_string(image)
        if text and text.strip():
            return text.strip()
        return "(Image processed but no text detected by OCR)"
    except ImportError:
        logger.warning("pytesseract not installed")
    except Exception as e:
        logger.warning("pytesseract OCR failed: %s", e)

    # Try easyocr as fallback
    try:
        import easyocr
        import numpy as np
        from PIL import Image

        file_storage.seek(0)
        image = Image.open(file_storage)
        img_array = np.array(image)

        reader = easyocr.Reader(['en', 'fr'], gpu=False)
        results = reader.readtext(img_array)
        text = "\n".join([r[1] for r in results])
        if text.strip():
            return text.strip()
        return "(Image processed but no text detected by OCR)"
    except ImportError:
        logger.warning("easyocr not installed")
    except Exception as e:
        logger.warning("easyocr OCR failed: %s", e)

    # Last resort: try with just Pillow to confirm it's a valid image
    try:
        from PIL import Image
        file_storage.seek(0)
        image = Image.open(file_storage)
        width, height = image.size
        return f"(Image detected: {width}x{height} pixels, format: {image.format}. OCR not available — install pytesseract: pip install pytesseract, and Tesseract-OCR system package)"
    except ImportError:
        return "(Image support requires Pillow: pip install Pillow)"
    except Exception:
        return "(Could not process image file)"


def _extract_excel(file_storage) -> str:
    """Extract text content from Excel files (.xlsx, .xls, .ods)."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
        lines = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    lines.append(row_text)

        wb.close()
        return "\n".join(lines) if lines else "(Excel file contained no data)"

    except ImportError:
        return "(Excel support requires openpyxl: pip install openpyxl)"
    except Exception as e:
        return f"(Excel extraction failed: {e})"
